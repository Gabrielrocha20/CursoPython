"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
"""

import openpyxl
from random import uniform
"""
pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

for linha in range(5,16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco
pedidos.save('nova_planilha.xlsx')"""

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha2.cell(linha, 1).value = f'Gabriel {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Joao {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Milena {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_pnl.xlsx')
